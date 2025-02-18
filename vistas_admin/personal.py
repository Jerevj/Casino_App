import random
import re
from tkinter import *
from tkinter import ttk, messagebox
import tkinter as tk
from conexion import Conexion  # Importar la clase Conexion
from utils.excel_utils import excel_manager  # Importar el gestor de Excel
from openpyxl import load_workbook

class Personal(tk.Frame):
    def __init__(self, padre):
        super().__init__(padre)
        self.db = Conexion()  # Instanciamos la conexión a la BD
        self.db.conectar()  # Establecemos la conexión
        self.estado = 1
        self.widgets()
        self.cargar_personal()  # Cargar los datos automáticamente

    def widgets(self):
        # Etiqueta de título
        tk.Label(self, text="Mantenedor de Personal", font=("Arial", 16)).grid(row=0, column=0, pady=10, columnspan=3)

        # Frame para la tabla
        self.tabla_frame = Frame(self)
        self.tabla_frame.grid(row=1, column=0, columnspan=2, padx=5, pady=5, sticky="nsew")

        # Crear el Treeview
        self.treeview = ttk.Treeview(self.tabla_frame, columns=("RUT", "Nombre", "Clave", "Estado"), show="headings")
        self.treeview.grid(row=0, column=0, sticky="nsew")

        # Configurar las columnas
        for col in ("RUT", "Nombre", "Clave", "Estado"):
            self.treeview.heading(col, text=col)
            self.treeview.column(col, width=200)

        # Scrollbar
        scrollbar = Scrollbar(self.tabla_frame, orient=VERTICAL, command=self.treeview.yview)
        scrollbar.grid(row=0, column=1, sticky="ns")
        self.treeview.config(yscrollcommand=scrollbar.set)

        # Permitir edición con doble clic
        self.treeview.bind("<Double-1>", self.editar_celda)

        # Frame para los botones
        self.botones_frame = Frame(self)
        self.botones_frame.grid(row=1, column=2, padx=10, pady=5, sticky="nsew")

        # Botón para actualizar la tabla
        tk.Button(self.botones_frame, text="🔄 Actualizar Personal", command=self.cargar_personal, width=20, height=2).grid(row=0, column=0, pady=5)

        # Botón para abrir el formulario de agregar empleado
        tk.Button(self.botones_frame, text="➕ Agregar Persona", command=self.abrir_formulario, width=20, height=2).grid(row=1, column=0, pady=5)

        # Botón para desactivar empleados
        tk.Button(self.botones_frame, text="❌ Desactivar Empleado", command=self.desactivar_empleado, fg="red", width=20, height=2).grid(row=2, column=0, pady=5)

        # Botón para ACTIVAR empleados
        tk.Button(self.botones_frame, text="✔️ Activar Empleado", command=self.activar_empleado, fg="blue", width=20, height=2).grid(row=3, column=0, pady=5)
        
        # Botón para alternar entre activos e inactivos
        self.boton_estado = tk.Button(self.botones_frame, text="Mostrar Inactivos", command=self.alternar_estado, width=20, height=2)
        self.boton_estado.grid(row=4, column=0, pady=5)

        # Botón para generar clave única
        tk.Button(self.botones_frame, text="🔑 Generar Clave Única", command=self.generar_clave_unica, width=20, height=2).grid(row=5, column=0, pady=5)

        # Hacer el diseño responsivo
        self.grid_rowconfigure(1, weight=1)
        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)
        self.grid_columnconfigure(2, weight=0)
        self.tabla_frame.grid_rowconfigure(0, weight=1)
        self.tabla_frame.grid_columnconfigure(0, weight=1)

    def cargar_personal(self):
        """Carga los datos de la tabla personas en el Treeview."""
        for widget in self.treeview.get_children():
            self.treeview.delete(widget)

        query = "SELECT rut, nombre, clave, estado FROM personas WHERE estado = %s"
        self.db.cursor.execute(query, (self.estado,))
        personas = self.db.cursor.fetchall()

        for persona in personas:
            self.treeview.insert("", "end", values=persona)

    def alternar_estado(self):
        """Alterna entre mostrar empleados activos e inactivos."""
        self.estado = 0 if self.estado == 1 else 1
        self.boton_estado.config(text="Mostrar Activos" if self.estado == 0 else "Mostrar Inactivos")
        self.cargar_personal()

    def editar_celda(self, event):
        """Permite editar una celda al hacer doble clic."""
        try:
            item = self.treeview.selection()[0]
        except IndexError:
            messagebox.showwarning("Atención", "Seleccione un elemento para editar.")
            return

        col = self.treeview.identify_column(event.x)
        col_idx = int(col[1:]) - 1  # Convertir a índice de columna

        old_value = self.treeview.item(item, "values")[col_idx]

        # Crear un Entry en la celda seleccionada
        entry = tk.Entry(self.treeview)
        entry.insert(0, old_value)
        entry.select_range(0, tk.END)
        entry.focus()

        def guardar_cambio(event=None):
            nuevo_valor = entry.get()
            valores = list(self.treeview.item(item, "values"))
            valores[col_idx] = nuevo_valor

            # Validaciones
            if col_idx == 0:  # Validar RUT
                if not re.match(r"^\d{7,8}-[0-9Kk]$", nuevo_valor):
                    messagebox.showwarning("Error", "El RUT debe tener el formato correcto (ej. 12345678-9).")
                    entry.destroy()
                    return
                query = "SELECT * FROM personas WHERE rut = %s"
                self.db.cursor.execute(query, (nuevo_valor,))
                if self.db.cursor.fetchone():
                    messagebox.showwarning("Error", "Ya existe un empleado con este RUT.")
                    entry.destroy()
                    return
                # Confirmación de cambio de RUT
                if not messagebox.askyesno("Confirmación", f"¿Está seguro de que desea cambiar el RUT a {nuevo_valor}?"):
                    entry.destroy()
                    return
            elif col_idx == 1:  # Convertir nombre a mayúsculas
                valores[col_idx] = nuevo_valor.upper()
            elif col_idx == 2:  # Validar clave
                if len(nuevo_valor) != 4 or not nuevo_valor.isdigit():
                    messagebox.showwarning("Error", "La clave debe ser de 4 dígitos numéricos.")
                    entry.destroy()
                    return
                query = "SELECT * FROM personas WHERE clave = %s"
                self.db.cursor.execute(query, (nuevo_valor,))
                if self.db.cursor.fetchone():
                    messagebox.showwarning("Error", "Ya existe un empleado con esta clave.")
                    entry.destroy()
                    return
            elif col_idx == 3:  # Validar estado
                if nuevo_valor not in ["0", "1"]:
                    messagebox.showwarning("Error", "El estado solo puede ser 0 o 1.")
                    entry.destroy()
                    return

            self.treeview.item(item, values=valores)

            # Actualizar en la base de datos
            rut = valores[0]
            campos = ["rut", "nombre", "clave", "estado"]
            campo_a_modificar = campos[col_idx]
            query = f"UPDATE personas SET {campo_a_modificar} = %s WHERE rut = %s"
            self.db.cursor.execute(query, (nuevo_valor, rut))
            self.db.conexion.commit()

            # Sincronizar cambios con el Excel
            self.sincronizar_persona_con_excel(rut, valores[1])  # Pasar el RUT y el nombre

            entry.destroy()

        entry.bind("<Return>", guardar_cambio)
        entry.bind("<FocusOut>", lambda e: entry.destroy())

        bbox = self.treeview.bbox(item, column=col)
        entry.place(x=bbox[0], y=bbox[1], width=bbox[2], height=bbox[3])

    def abrir_formulario(self):
        """Abre una nueva ventana para agregar un empleado."""
        formulario = Toplevel(self)
        formulario.title("Agregar Empleado")
        formulario.geometry("300x250")
        formulario.resizable(False, False)
    
        # Títulos de los campos
        tk.Label(formulario, text="RUT:").grid(row=0, column=0, pady=5, padx=10, sticky=W)
        rut_entry = tk.Entry(formulario)
        rut_entry.grid(row=0, column=1, pady=5, padx=10)

        tk.Label(formulario, text="Nombre:").grid(row=1, column=0, pady=5, padx=10, sticky=W)
        nombre_entry = tk.Entry(formulario)
        nombre_entry.grid(row=1, column=1, pady=5, padx=10)

        tk.Label(formulario, text="Clave:").grid(row=2, column=0, pady=5, padx=10, sticky=W)
        clave_entry = tk.Entry(formulario)
        clave_entry.grid(row=2, column=1, pady=5, padx=10)

        def agregar_empleado():
            """Agrega un nuevo empleado a la base de datos desde el formulario."""
            rut = rut_entry.get()
            nombre = nombre_entry.get().upper()  # Convertir nombre a mayúsculas
            clave = clave_entry.get()

            # Validar formato del RUT (ej. 12345678-9 o 12.345.678-9)
            if not re.match(r"^\d{7,8}-[0-9Kk]$", rut):
                messagebox.showwarning("Error", "El RUT debe tener el formato correcto (ej. 12345678-9).")
                return

            # Validar que la clave tenga 4 dígitos
            if len(clave) != 4 or not clave.isdigit():
                messagebox.showwarning("Error", "La clave debe ser de 4 dígitos numéricos.")
                return

            # Verificar que la clave sea única
            query = "SELECT * FROM personas WHERE clave = %s"
            self.db.cursor.execute(query, (clave,))
            if self.db.cursor.fetchone():
                messagebox.showwarning("Error", "Ya existe un empleado con esta clave.")
                return

            if not rut or not nombre or not clave:
                messagebox.showwarning("Error", "Todos los campos son obligatorios.")
                return

            # Formatear el nombre
            nombre_parts = nombre.split()
            if len(nombre_parts) == 4:
                nombre_formateado = f"{nombre_parts[2].upper()} {nombre_parts[3].upper()} {nombre_parts[0].upper()} {nombre_parts[1].upper()}"
            elif len(nombre_parts) == 3:
                nombre_formateado = f"{nombre_parts[1].upper()} {nombre_parts[2].upper()} {nombre_parts[0].upper()}"
            else:
                nombre_formateado = nombre.upper()  # En caso de que el nombre no tenga el formato esperado

            query = "INSERT INTO personas (rut, nombre, clave, estado) VALUES (%s, %s, %s, %s)"
            try:
                self.db.cursor.execute(query, (rut, nombre_formateado, clave, 1))
                self.db.conexion.commit()
                self.sincronizar_persona_con_excel(rut, nombre_formateado)  # Sincronizar con Excel
                self.cargar_personal()
                messagebox.showinfo("Éxito", "Empleado agregado correctamente.")
                formulario.destroy()  # Cerrar el formulario
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo agregar el empleado: {e}")

        # Botón para agregar empleado
        tk.Button(formulario, text="Agregar", command=agregar_empleado, width=20).grid(row=3, column=0, columnspan=2, pady=10)

    def sincronizar_persona_con_excel(self, rut, nombre, desactivar=False):
        try:
            minuta_path = excel_manager.obtener_ruta_minuta()
            wb = load_workbook(minuta_path)
            sheet = wb.active

            # Verificar si el RUT ya está en el Excel
            for row in range(2, sheet.max_row + 1):
                if sheet.cell(row=row, column=1).value == rut:
                    if desactivar:
                        sheet.cell(row=row, column=34, value=0)  # Columna 33 (Estado) a 0 (Inactivo)
                        print(f"Empleado {rut} desactivado en el Excel.")  # Mensaje de depuración
                    else:
                        sheet.cell(row=row, column=33, value=nombre)  # Actualizar el nombre
                    break
            else:
                if not desactivar:
                    fila_actual = sheet.max_row + 1
                    sheet.cell(row=fila_actual, column=1, value=rut)  # Columna 0 (RUT)
                    sheet.cell(row=fila_actual, column=33, value=nombre)  # Columna 32 (Nombre Funcionario)

            wb.save(minuta_path)
            wb.close()  # Cerrar el archivo después de guardar
        except Exception as e:
            messagebox.showerror("Error", f"Error al sincronizar persona con el Excel: {e}")

    def desactivar_empleado(self):
        """Cambia el estado de un empleado a 'Inactivo'."""
        try:
            item = self.treeview.selection()[0]
            rut = self.treeview.item(item, "values")[0]
            query = "UPDATE personas SET estado = 0 WHERE rut = %s"
            self.db.cursor.execute(query, (rut,))
            self.db.conexion.commit()
            print(f"Empleado {rut} desactivado en la base de datos.")  # Mensaje de depuración
            self.sincronizar_persona_con_excel(rut, None, desactivar=True)  # Sincronizar con Excel
            self.cargar_personal()
            messagebox.showinfo("Éxito", "Empleado desactivado correctamente.")
        except IndexError:
            messagebox.showwarning("Atención", "Seleccione un empleado para desactivar.")
        except Exception as e:
            messagebox.showerror("Error", f"Error al desactivar el empleado: {e}")

    def activar_empleado(self):
        """Cambia el estado de un empleado a 'Activo'."""
        try:
            item = self.treeview.selection()[0]
            rut = self.treeview.item(item, "values")[0]
            query = "UPDATE personas SET estado = 1 WHERE rut = %s"
            self.db.cursor.execute(query, (rut,))
            self.db.conexion.commit()
            print(f"Empleado {rut} Activado en la base de datos.")  # Mensaje de depuración
            self.sincronizar_persona_con_excel(rut, None, desactivar=False)  # Sincronizar con Excel
            self.cargar_personal()
            messagebox.showinfo("Éxito", "Empleado activado correctamente.")
        except IndexError:
            messagebox.showwarning("Atención", "Seleccione un empleado para activar.")
        except Exception as e:
            messagebox.showerror("Error", f"Error al activar el empleado: {e}")

    def generar_clave_unica(self):
        """Genera una clave única de 4 dígitos que no esté repetida en la base de datos y la copia al portapapeles."""
        while True:
            clave = f"{random.randint(0, 9999):04d}"
            query = "SELECT * FROM personas WHERE clave = %s"
            self.db.cursor.execute(query, (clave,))
            if not self.db.cursor.fetchone():
                self.clipboard_clear()
                self.clipboard_append(clave)
                messagebox.showinfo("Clave Generada", f"La clave generada es: {clave}\nLa clave ha sido copiada al portapapeles.")
                break

    def on_closing(self):
        """Cierra la conexión cuando la ventana se cierra."""
        self.db.desconectar()
        self.master.destroy()

# Ejemplo de uso
if __name__ == "__main__":
    root = Tk()
    root.title("Mantenedor de Personal")
    root.geometry("800x500")
    root.resizable(True, True)  # Permitir que la ventana cambie de tamaño
    personal_frame = Personal(root)
    personal_frame.pack(fill=BOTH, expand=True)
    root.protocol("WM_DELETE_WINDOW", personal_frame.on_closing)
    root.mainloop()