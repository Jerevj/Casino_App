import tkinter as tk
from tkinter import ttk, messagebox
from utils.excel_utils import excel_manager
from openpyxl import load_workbook

class Minutas(tk.Frame):
    def __init__(self, padre, db_connection):
        super().__init__(padre)
        self.db_connection = db_connection  # Guardar la conexión
        self.data = []
        self.column_names = ['Día', 'Menú A', 'Menú B', 'Menú C']
        self.widgets()
        self.cargar_minutas()  # Llamar a cargar_minutas aquí

    def widgets(self):
        titulo = tk.Label(self, text="Gestión de Menús seleccionados", font=("Arial", 20, "bold"))
        titulo.pack(pady=10)
        
        self.btn_actualizar = tk.Button(self, text="Actualizar", command=self.actualizar_datos)
        self.btn_actualizar.pack(pady=5)
        
        self.main_frame = tk.Frame(self)
        self.main_frame.pack(fill=tk.BOTH, expand=True)
        
        self.scrollbar_vertical = ttk.Scrollbar(self.main_frame, orient=tk.VERTICAL)
        self.scrollbar_vertical.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.scrollbar_horizontal = ttk.Scrollbar(self, orient=tk.HORIZONTAL)
        self.scrollbar_horizontal.pack(side=tk.BOTTOM, fill=tk.X)
        
        self.table_frame = tk.Frame(self.main_frame)
        self.table_frame.pack(fill=tk.BOTH, expand=True)

    def actualizar_datos(self):
        self.sincronizar_empleados_con_excel()
        self.recargar_y_actualizar()

    def sincronizar_empleados_con_excel(self):
        try:
            print("Sincronizando empleados con Excel...")  # Mensaje de depuración
            minuta_path = excel_manager.obtener_ruta_minuta()
            wb = load_workbook(minuta_path)
            sheet = wb.active
            
            # Obtener lista de RUTs en Excel
            ruts_excel = {sheet.cell(row=i, column=1).value for i in range(2, sheet.max_row + 1)}
            
            # Obtener lista de empleados activos desde la BD
            self.db_connection.cursor.execute("SELECT rut, nombre FROM personas WHERE estado = 1")
            empleados = self.db_connection.cursor.fetchall()
            
            fila_actual = sheet.max_row + 1
            for rut, nombre in empleados:
                if rut not in ruts_excel:
                    sheet.cell(row=fila_actual, column=1, value=rut)  # Columna 0 (RUT)
                    sheet.cell(row=fila_actual, column=33, value=nombre)  # Columna 32 (Nombre Funcionario)
                    fila_actual += 1
                else:
                    # Actualizar el nombre si el RUT ya existe
                    for row in range(2, sheet.max_row + 1):
                        if sheet.cell(row=row, column=1).value == rut:
                            sheet.cell(row=row, column=33, value=nombre)  # Actualizar el nombre

            wb.save(minuta_path)
            wb.close()  # Cerrar el archivo después de guardar
        
            messagebox.showinfo("Éxito", "Los empleados fueron sincronizados con el Excel.")
        except Exception as e:
            messagebox.showerror("Error", f"Error al sincronizar empleados con el Excel: {e}")
        
    def cargar_minutas(self):
        try:
            print("Cargando minutas...")  # Mensaje de depuración
            sheet = excel_manager.obtener_minuta_sheet()
            if not sheet:
                raise Exception("No se pudo obtener la hoja de minutas.")

            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                raise Exception("El archivo está vacío o mal estructurado.")

            # Validar que el archivo tenga al menos 33 columnas
            if len(rows[0]) < 33:
                raise Exception("El archivo Excel no tiene el formato esperado. Asegúrate de que tenga al menos 33 columnas.")

            self.data = []
            self.column_names = rows[0]

            for widget in self.table_frame.winfo_children():
                widget.destroy()

            self.tree = ttk.Treeview(self.table_frame, show="headings", columns=self.column_names, height=25)
            self.tree.pack(fill=tk.BOTH, expand=True)

            self.scrollbar_vertical.config(command=self.tree.yview)
            self.scrollbar_horizontal.config(command=self.tree.xview)
            self.tree.configure(yscrollcommand=self.scrollbar_vertical.set, xscrollcommand=self.scrollbar_horizontal.set)

            for col in self.column_names:
                self.tree.heading(col, text=col)
                self.tree.column(col, anchor="center", width=130)
            
            ruts_vistos = set()
            for row in rows[1:]:
                rut = row[0]
                if rut in ruts_vistos:
                    continue  # Saltar filas duplicadas
                ruts_vistos.add(rut)
                # Verificar si el RUT está activo en la base de datos
                self.db_connection.cursor.execute("SELECT estado FROM personas WHERE rut = %s", (rut,))
                estado = self.db_connection.cursor.fetchone()
                if estado and estado[0] == 1:  # Solo mostrar si el estado es activo (1)
                    self.tree.insert('', 'end', values=row)
                else:
                    print(f"Empleado inactivo: {rut}")  # Mensaje de depuración
            
            self.tree.bind("<Double-1>", self.editar_celda)

        except Exception as e:
            messagebox.showerror("Error", f"Error al cargar las minutas desde el archivo: {e}")       
        

    def editar_celda(self, event):
        item = self.tree.identify_row(event.y)
        column = self.tree.identify_column(event.x)
        col_index = int(column[1:]) - 1  # Convertir el índice a entero

        # Bloquear edición en la primera columna (RUT) y la columna 32 (Nombre Funcionario)
        if col_index == 0 or col_index == 32:
            return  # No permitir edición en estas columnas
    
        if item:
            valores = list(self.tree.item(item, "values"))
            valor_actual = valores[col_index]

            entrada = tk.Entry(self.tree)
            entrada.insert(0, valor_actual)
            entrada.bind("<Return>", lambda e: self.guardar_edicion(item, col_index, entrada.get()))
            entrada.bind("<FocusOut>", lambda e: entrada.destroy())

            x, y, width, height = self.tree.bbox(item, column)
            entrada.place(x=x, y=y, width=width, height=height)
            entrada.focus()

    def guardar_edicion(self, item, col_index, nuevo_valor):
        valores = list(self.tree.item(item, "values"))
        valores[col_index] = nuevo_valor
        self.tree.item(item, values=valores)
        
        self.actualizar_excel()
        self.recargar_y_actualizar()  # Recargar y actualizar después de actualizar el Excel
    
    def actualizar_excel(self):
        try:
            print("Actualizando Excel...")  # Mensaje de depuración
            minuta_path = excel_manager.obtener_ruta_minuta()
            wb = load_workbook(minuta_path)
            sheet = wb.active
            for i, item in enumerate(self.tree.get_children(), start=2):
                valores = self.tree.item(item, "values")
                for j, valor in enumerate(valores):
                    sheet.cell(row=i, column=j+1, value=valor)
            wb.save(minuta_path)
            wb.close()  # Cerrar el archivo después de guardar
            messagebox.showinfo("Éxito", "Datos actualizados en Excel.")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo actualizar el Excel: {e}")

    def recargar_y_actualizar(self):
        """Recarga el archivo Excel y actualiza el Treeview."""
        print("Recargando archivos Excel...")  # Mensaje de depuración
        excel_manager.recargar_archivos()  # Recargar los archivos Excel
        self.cargar_minutas()  # Actualizar el Treeview