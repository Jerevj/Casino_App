# Clase ExcelManager
import os
from openpyxl import load_workbook
from tkinter import messagebox
from config import MINUTA_FILE_PATH, MENUS_FILE_PATH

class ExcelManager:
    def __init__(self):
        if not os.path.exists(MINUTA_FILE_PATH):
            messagebox.showerror("Error", f"No se encontró el archivo: {MINUTA_FILE_PATH}")
            return

        if not os.path.exists(MENUS_FILE_PATH):
            messagebox.showerror("Error", f"No se encontró el archivo: {MENUS_FILE_PATH}")
            return

        self.minuta_file_path = MINUTA_FILE_PATH
        self.menus_file_path = MENUS_FILE_PATH
        self.minuta_wb = None
        self.menus_wb = None
        self.minuta_ws = None
        self.menus_ws = None
        self.cargar_archivos()

    def cargar_archivos(self):
        """Carga los archivos de Excel usando openpyxl."""
        try:
            self.minuta_wb = load_workbook(self.minuta_file_path)
            self.minuta_ws = self.minuta_wb.active  # Asumimos que los datos están en la primera hoja
        except Exception as e:
            print(f"Error al cargar 'Minuta_Actual.xlsx': {e}")
            messagebox.showerror("Error", "No se pudo cargar el archivo 'Minuta_Actual.xlsx'.")
            self.minuta_w
        try:
            self.menus_wb = load_workbook(self.menus_file_path)
            self.menus_ws = self.menus_wb.active  # Cargar la hoja de 'Menus_Actual.xlsx'
        except Exception as e:
            print(f"Error al cargar 'Menus_Actual.xlsx': {e}")
            messagebox.showerror("Error", "No se pudo cargar el archivo 'Menus_Actual.xlsx'.")
            self.menus_ws = None  # Asegurar que no haya una referencia incorrecta

    def recargar_archivos(self):
        """Recarga los archivos de Excel para reflejar los cambios."""
        self.cargar_archivos()
    def obtener_menu_desde_excel(self, rut, dia):
        """Obtiene la opción de menú (A, B, C) de un empleado según el RUT y el día del mes desde 'Minuta_Actual.xlsx'."""
        if self.minuta_ws is None:
            print("El archivo 'Minuta_Actual.xlsx' no está cargado correctamente.")
            return None
        
        # Buscar la fila con el RUT correspondiente
        for row in self.minuta_ws.iter_rows(min_row=2, values_only=True):  # Asumimos que la primera fila es el encabezado
            if str(row[0]) == str(rut):  # Comparar el RUT
                columna_dia = str(dia)  # Asegurar que el día sea string
                # Buscar el día en los encabezados
                for col_num, cell in enumerate(list(self.minuta_ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]):
                    if str(cell) == columna_dia:  # Buscar el día en los encabezados
                        menu = row[col_num]  # Obtener la opción A, B o C
                        if menu:
                            return menu
                        else:
                            print(f"No hay menú asignado para el RUT {rut} el día {dia}.")
                            return None
        print(f"RUT {rut} no encontrado en el Excel.")
        return None

    def obtener_nombre_menu(self, dia, opcion_menu):
        """Obtiene el nombre del menú a partir de la opción (A, B o C) desde 'Menus_Actual.xlsx'."""
        if self.menus_ws is None:
            print("El archivo 'Menus_Actual.xlsx' no está cargado correctamente.")
            return None
        
        # Iterar sobre las filas (asumiendo que la primera fila es el encabezado)
        for row in self.menus_ws.iter_rows(min_row=2, values_only=True):
            fecha_cell = row[0]
            # Si la celda está vacía, saltar sin imprimir mensajes
            if not fecha_cell:
                continue

            fecha = str(fecha_cell)
            
            # Verificar que la fecha tenga exactamente 8 caracteres (formato yyyymmdd)
            if len(fecha) != 8:
                # Podrías imprimir un mensaje de depuración una sola vez o simplemente omitirlo
                # print(f"Fecha inválida en la fila: {fecha}. Saltando esta fila.")
                continue

            try:
                # Extraer los dos últimos dígitos que corresponden al día del mes
                dia_mes_menu = int(fecha[-2:])
            except ValueError:
                # Si ocurre un error al convertir, saltamos la fila
                # print(f"Fecha no válida en la fila: {fecha}. No se puede extraer el día.")
                continue

            # Comparar el día del mes con el día buscado
            if dia_mes_menu == dia:
                # Dependiendo de la opción de menú (A, B o C) se devuelve la columna correspondiente
                if opcion_menu == 'A':
                    return row[1]  # Menú A en la segunda columna
                elif opcion_menu == 'B':
                    return row[2]  # Menú B en la tercera columna
                elif opcion_menu == 'C':
                    return row[3]  # Menú C en la cuarta columna
                else:
                    print(f"Opción de menú {opcion_menu} no válida.")
                    return "Opción de menú no válida"

        # Si se terminó el ciclo sin encontrar una fila que cumpla, se informa que no hay menú
        print(f"No hay menú para el día {dia} con la opción {opcion_menu}.")
        return "No hay menú disponible para este día."

    def obtener_minuta_sheet(self):
        """Devuelve la hoja activa del archivo 'Minuta_Actual.xlsx' si está cargado correctamente."""
        if self.minuta_ws is None:
            messagebox.showerror("Error", "No se pudo obtener la hoja de Minuta_Actual.xlsx. Asegúrate de que el archivo existe y está bien estructurado.")
            return None
        return self.minuta_ws

    def obtener_ruta_minuta(self):
        return self.minuta_file_path
    
    def obtener_ruta_menus(self):
        return self.menus_file_path

# Crear la instancia global de ExcelManager
excel_manager = ExcelManager()